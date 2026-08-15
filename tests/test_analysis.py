from __future__ import annotations

import json
import unittest
from pathlib import Path

from openpyxl import load_workbook

from backend.app.repository import DashboardRepository


ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data" / "processed" / "dashboard.json"
EXPORTS = ROOT / "web" / "public" / "exports"


class DetailedAnalysisTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.dashboard = DashboardRepository(DATA).load()
        cls.raw = json.loads(DATA.read_text(encoding="utf-8"))

    def test_contract_contains_both_models(self) -> None:
        self.assertEqual(self.dashboard.schema_version, "5.0")
        regression = self.dashboard.analysis.regression
        clustering = self.dashboard.analysis.clustering
        self.assertGreater(regression.sample_size, 500)
        self.assertEqual(len(regression.coefficients), 4)
        self.assertEqual(clustering.selected_k, len(clustering.clusters))
        self.assertEqual(
            clustering.sample_size,
            sum(cluster.zip_count for cluster in clustering.clusters),
        )

    def test_regression_metrics_are_cross_validated_and_consistent(self) -> None:
        regression = self.dashboard.analysis.regression
        self.assertGreater(regression.r2, 0.5)
        self.assertLess(regression.r2, 1)
        self.assertGreater(regression.mae, 0)
        self.assertEqual(regression.sample_size, len(regression.predictions))
        self.assertIn("5 katlı", regression.notes[2])

    def test_clustering_has_unique_assignments_and_k_evaluation(self) -> None:
        clustering = self.dashboard.analysis.clustering
        zips = [assignment.zip_code for assignment in clustering.assignments]
        self.assertEqual(len(zips), len(set(zips)))
        self.assertEqual({row.k for row in clustering.k_evaluation}, set(range(2, 7)))
        selected = next(
            row for row in clustering.k_evaluation if row.k == clustering.selected_k
        )
        self.assertEqual(selected.silhouette, clustering.silhouette_score)

    def test_excel_exports_have_charts_formulas_and_method_sheets(self) -> None:
        regression_path = EXPORTS / "regression_analysis.xlsx"
        clustering_path = EXPORTS / "clustering_analysis.xlsx"
        self.assertTrue(regression_path.exists())
        self.assertTrue(clustering_path.exists())

        regression = load_workbook(regression_path, data_only=False)
        self.assertEqual(
            regression.sheetnames,
            ["Özet", "Katsayılar", "Tahminler", "Yöntem"],
        )
        self.assertGreaterEqual(len(regression["Özet"]._charts), 1)
        self.assertGreaterEqual(len(regression["Katsayılar"]._charts), 1)
        self.assertTrue(str(regression["Özet"]["C12"].value).startswith("="))
        self.assertTrue(str(regression["Tahminler"]["F2"].value).startswith("="))
        self.assertEqual(regression["Tahminler"]["D2"].alignment.horizontal, "center")

        clustering = load_workbook(clustering_path, data_only=False)
        self.assertEqual(
            clustering.sheetnames,
            ["Özet", "Küme Profilleri", "K Seçimi", "ZIP Atamaları", "Yöntem"],
        )
        self.assertGreaterEqual(len(clustering["Küme Profilleri"]._charts), 1)
        self.assertGreaterEqual(len(clustering["K Seçimi"]._charts), 1)
        self.assertEqual(clustering["ZIP Atamaları"]["J2"].alignment.horizontal, "center")


if __name__ == "__main__":
    unittest.main()
