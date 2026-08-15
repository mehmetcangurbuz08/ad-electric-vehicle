from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference, ScatterChart, Series
from openpyxl.cell.cell import MergedCell
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

from data.pipeline.modeling import ModelTables


NAVY = "17324D"
BLUE = "2F6FED"
GREEN = "2F9E72"
ORANGE = "EF8354"
LIGHT = "EAF0F6"
MUTED = "60758A"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="D9E2EC")


def _title(sheet, title: str, subtitle: str, end_column: int = 8) -> None:
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    cell = sheet.cell(1, 1, title)
    cell.font = Font(size=20, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 34
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_column)
    sheet.cell(2, 1, subtitle).font = Font(size=10, color=MUTED, italic=True)
    sheet.row_dimensions[2].height = 25
    sheet.freeze_panes = "A4"
    sheet.sheet_view.showGridLines = False


def _section(sheet, row: int, text: str, end_column: int = 8) -> None:
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_column)
    cell = sheet.cell(row, 1, text)
    cell.font = Font(size=12, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=BLUE)
    cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[row].height = 23


def _headers(sheet, row: int, headers: list[str]) -> None:
    for column, value in enumerate(headers, 1):
        cell = sheet.cell(row, column, value)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN)
    sheet.row_dimensions[row].height = 28


def _table(sheet, reference: str, name: str) -> None:
    table = Table(displayName=name, ref=reference)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def _widths(sheet, widths: dict[str, float]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def _center_workbook_cells(workbook: Workbook) -> None:
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell) or cell.value is None:
                    continue
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=cell.alignment.wrap_text,
                    shrink_to_fit=cell.alignment.shrink_to_fit,
                )


def _method_sheet(workbook: Workbook, title: str, rows: list[tuple[str, str]]) -> None:
    sheet = workbook.create_sheet("Yöntem")
    _title(sheet, title, "Formüller, veri hazırlığı ve yorumlama notları", 6)
    _headers(sheet, 4, ["Başlık", "Açıklama"])
    for row_index, (heading, detail) in enumerate(rows, 5):
        sheet.cell(row_index, 1, heading).font = Font(bold=True, color=NAVY)
        sheet.cell(row_index, 2, detail).alignment = Alignment(wrap_text=True, vertical="top")
        sheet.cell(row_index, 1).alignment = Alignment(vertical="top")
        sheet.row_dimensions[row_index].height = max(34, min(90, len(detail) // 2))
    _widths(sheet, {"A": 28, "B": 100})
    sheet.freeze_panes = "A5"


def _regression_workbook(
    destination: Path, analysis: dict, tables: ModelTables
) -> None:
    workbook = Workbook()
    workbook.calculation.fullCalcOnLoad = True
    summary = workbook.active
    summary.title = "Özet"
    _title(
        summary,
        "Çoklu Doğrusal Regresyon",
        "Washington ZIP/ZCTA bölgelerinde EV yoğunluğunu açıklayan Census göstergeleri",
        9,
    )
    _section(summary, 4, "Model tanımı", 9)
    definitions = [
        ("Hedef", analysis["target"]),
        ("Örneklem", analysis["sampleSize"]),
        ("Doğrulama", "5 katlı çapraz doğrulama; her ZIP eğitim dışında tahmin edildi"),
        ("Denklem", analysis["formula"]),
    ]
    for row_index, (label, value) in enumerate(definitions, 5):
        summary.cell(row_index, 1, label).font = Font(bold=True, color=NAVY)
        summary.cell(row_index, 2, value)
        summary.merge_cells(start_row=row_index, start_column=2, end_row=row_index, end_column=9)
        summary.cell(row_index, 2).alignment = Alignment(wrap_text=True)

    _section(summary, 10, "Başarı ölçüleri", 9)
    _headers(summary, 11, ["Ölçü", "Python sonucu", "Excel kontrolü", "Nasıl okunur?"])
    last_prediction_row = len(tables.regression_predictions) + 1
    metric_rows = [
        (
            "R²",
            analysis["r2"],
            f"=1-SUM('Tahminler'!H2:H{last_prediction_row})/DEVSQ('Tahminler'!D2:D{last_prediction_row})",
            "Modelin ZIP'ler arasındaki EV yoğunluğu farkının ne kadarını açıkladığını gösterir; yüksek olması iyidir.",
        ),
        (
            "MAE",
            analysis["mae"],
            f"=AVERAGE('Tahminler'!G2:G{last_prediction_row})",
            "Tahminlerin gerçek değerden ortalama mutlak sapmasıdır; düşük olması iyidir.",
        ),
        (
            "RMSE",
            analysis["rmse"],
            f"=SQRT(AVERAGE('Tahminler'!H2:H{last_prediction_row}))",
            "Büyük hatalara MAE'den daha fazla ağırlık verir; düşük olması iyidir.",
        ),
        (
            "CV R² ortalaması",
            analysis["cvR2Mean"],
            "Python 5-fold sonucu",
            "Beş doğrulama grubundaki R² değerlerinin ortalamasıdır.",
        ),
    ]
    for row_index, values in enumerate(metric_rows, 12):
        for column, value in enumerate(values, 1):
            summary.cell(row_index, column, value)
        summary.cell(row_index, 1).font = Font(bold=True, color=NAVY)
        summary.cell(row_index, 4).alignment = Alignment(wrap_text=True)
    _widths(summary, {"A": 23, "B": 18, "C": 24, "D": 78})

    coefficients = workbook.create_sheet("Katsayılar")
    _title(
        coefficients,
        "Standartlaştırılmış Regresyon Katsayıları",
        "+1 standart sapmalık değişimin EV/1.000 konut tahminindeki karşılığı",
        7,
    )
    coefficient_headers = [
        "Değişken", "Katsayı", "Yön", "Ortalama", "Standart sapma", "Yorum"
    ]
    _headers(coefficients, 4, coefficient_headers)
    for row_index, row in enumerate(analysis["coefficients"], 5):
        values = [
            row["label"], row["coefficient"], row["direction"], row["mean"],
            row["standardDeviation"], row["interpretation"],
        ]
        for column, value in enumerate(values, 1):
            coefficients.cell(row_index, column, value)
        coefficients.cell(row_index, 6).alignment = Alignment(wrap_text=True)
    coefficient_end = len(analysis["coefficients"]) + 4
    _table(coefficients, f"A4:F{coefficient_end}", "RegressionCoefficients")
    _widths(coefficients, {"A": 32, "B": 14, "C": 13, "D": 17, "E": 18, "F": 72})
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "Katsayı büyüklükleri"
    chart.y_axis.title = "Değişken"
    chart.x_axis.title = "EV / 1.000 konut"
    chart.add_data(Reference(coefficients, min_col=2, min_row=4, max_row=coefficient_end), titles_from_data=True)
    chart.set_categories(Reference(coefficients, min_col=1, min_row=5, max_row=coefficient_end))
    chart.height = 8
    chart.width = 15
    coefficients.add_chart(chart, "H4")

    predictions = workbook.create_sheet("Tahminler")
    prediction_headers = [
        "ZIP", "Şehir", "County", "Gerçek EV/1K", "Tahmin EV/1K",
        "Artık (Gerçek-Tahmin)", "Mutlak hata", "Hata karesi",
        "Medyan gelir", "Çok birimli konut %", "Evden çalışma %", "İşe gidiş dk",
    ]
    _headers(predictions, 1, prediction_headers)
    for row_index, row in enumerate(tables.regression_predictions.itertuples(), 2):
        values = [
            str(row.zipCode), row.city, row.county, row.evPer1kHousing,
            row.predictedEvPer1kHousing,
        ]
        for column, value in enumerate(values, 1):
            predictions.cell(row_index, column, value)
        predictions.cell(row_index, 6, f"=D{row_index}-E{row_index}")
        predictions.cell(row_index, 7, f"=ABS(F{row_index})")
        predictions.cell(row_index, 8, f"=F{row_index}^2")
        trailing = [
            row.medianIncome, row.multifamilyShare, row.workFromHomeShare,
            row.avgCommuteMinutes,
        ]
        for column, value in enumerate(trailing, 9):
            predictions.cell(row_index, column, value)
    _table(predictions, f"A1:L{last_prediction_row}", "RegressionPredictions")
    predictions.freeze_panes = "D2"
    predictions.auto_filter.ref = f"A1:L{last_prediction_row}"
    _widths(
        predictions,
        {"A": 11, "B": 20, "C": 18, "D": 15, "E": 15, "F": 21,
         "G": 14, "H": 14, "I": 17, "J": 20, "K": 18, "L": 15},
    )
    predictions.conditional_formatting.add(
        f"G2:G{last_prediction_row}",
        ColorScaleRule(start_type="min", start_color="E8F5E9", mid_type="percentile", mid_value=50, mid_color="FFF3CD", end_type="max", end_color="F8D7DA"),
    )
    scatter = ScatterChart()
    scatter.title = "Gerçek ve çapraz doğrulama tahminleri"
    scatter.x_axis.title = "Gerçek EV / 1.000 konut"
    scatter.y_axis.title = "Tahmin EV / 1.000 konut"
    series = Series(
        Reference(predictions, min_col=5, min_row=2, max_row=last_prediction_row),
        Reference(predictions, min_col=4, min_row=2, max_row=last_prediction_row),
        title="ZIP bölgeleri",
    )
    series.marker.symbol = "circle"
    series.marker.size = 4
    series.graphicalProperties.line.noFill = True
    scatter.series.append(series)
    scatter.height = 10
    scatter.width = 17
    summary.add_chart(scatter, "F11")

    _method_sheet(
        workbook,
        "Regresyon Yöntemi",
        [
            ("Araştırma sorusu", "Gelir, konut yapısı ve işe gidiş göstergeleri ZIP düzeyindeki EV yoğunluğunu ne ölçüde açıklıyor?"),
            ("Bağımlı değişken", "Y = (Kayıtlı EV / Konut birimi) × 1.000"),
            ("Model", "ŷ = β₀ + β₁z(Gelir) + β₂z(Çok birimli konut) + β₃z(Evden çalışma) + β₄z(İşe gidiş)"),
            ("Standardizasyon", "z = (x − ortalama) / standart sapma. Böylece farklı ölçü birimlerindeki katsayılar karşılaştırılabilir."),
            ("R²", "R² = 1 − Σ(y−ŷ)² / Σ(y−ȳ)². Açıklanan değişim oranıdır; yüksek değer daha güçlü açıklama anlamına gelir."),
            ("MAE", "MAE = Σ|y−ŷ| / n. Tahminin gerçek değerden ortalama kaç EV/1.000 konut saptığını gösterir."),
            ("Doğrulama", "5-fold çapraz doğrulamada veri beş gruba ayrılır; her grup, diğer dört grupla kurulan model tarafından tahmin edilir."),
            ("Sınır", "Kesitsel gözlemsel veri kullanıldığı için katsayılar nedensellik veya gelecek tahmini olarak yorumlanamaz."),
        ],
    )
    _center_workbook_cells(workbook)
    workbook.save(destination)


def _clustering_workbook(
    destination: Path, analysis: dict, tables: ModelTables
) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Özet"
    _title(
        summary,
        "K-Means Kümeleme Analizi",
        "EV yoğunluğu, kamu şarj portu yoğunluğu ve gelirle oluşturulan ZIP/ZCTA profilleri",
        9,
    )
    _section(summary, 4, "Model özeti", 9)
    summary_rows = [
        ("Kullanılan ZIP", analysis["sampleSize"]),
        ("Seçilen küme sayısı", analysis["selectedK"]),
        ("Silhouette score", analysis["silhouetteScore"]),
        ("Amaç fonksiyonu", analysis["formula"]),
    ]
    for row_index, (label, value) in enumerate(summary_rows, 5):
        summary.cell(row_index, 1, label).font = Font(bold=True, color=NAVY)
        summary.cell(row_index, 2, value)
        summary.merge_cells(start_row=row_index, start_column=2, end_row=row_index, end_column=7)
    _section(summary, 10, "Küme profilleri", 9)
    profile_headers = [
        "Küme", "Ad", "ZIP", "EV/1K konut", "Port/1K konut", "Medyan gelir",
        "Çok birimli %", "İşe gidiş dk", "Açıklama",
    ]
    _headers(summary, 11, profile_headers)
    for row_index, row in enumerate(analysis["clusters"], 12):
        values = [
            row["clusterId"], row["label"], row["zipCount"], row["evPer1kHousing"],
            row["portsPer1kHousing"], row["medianIncome"], row["multifamilyShare"],
            row["avgCommuteMinutes"], row["description"],
        ]
        for column, value in enumerate(values, 1):
            summary.cell(row_index, column, value)
        summary.cell(row_index, 1).fill = PatternFill("solid", fgColor=row["color"].lstrip("#"))
        summary.cell(row_index, 1).font = Font(bold=True, color=WHITE)
        summary.cell(row_index, 9).alignment = Alignment(wrap_text=True)
    profile_end = len(analysis["clusters"]) + 11
    _table(summary, f"A11:I{profile_end}", "ClusterSummary")
    _widths(summary, {"A": 10, "B": 29, "C": 10, "D": 16, "E": 17, "F": 18, "G": 17, "H": 16, "I": 70})

    profile_sheet = workbook.create_sheet("Küme Profilleri")
    _title(profile_sheet, "Küme Profilleri", "Küme ortalamaları ve toplamları", 12)
    profile_columns = [
        ("clusterId", "Küme"), ("label", "Ad"), ("zipCount", "ZIP sayısı"),
        ("evPer1kHousing", "EV/1K konut"), ("portsPer1kHousing", "Port/1K konut"),
        ("medianIncome", "Medyan gelir"), ("multifamilyShare", "Çok birimli %"),
        ("workFromHomeShare", "Evden çalışma %"), ("avgCommuteMinutes", "İşe gidiş dk"),
        ("bevShare", "BEV %"), ("vehicles", "Toplam EV"), ("publicPorts", "Toplam port"),
    ]
    _headers(profile_sheet, 4, [label for _, label in profile_columns])
    for row_index, row in enumerate(analysis["clusters"], 5):
        for column, (key, _) in enumerate(profile_columns, 1):
            profile_sheet.cell(row_index, column, row[key])
        profile_sheet.cell(row_index, 1).fill = PatternFill("solid", fgColor=row["color"].lstrip("#"))
        profile_sheet.cell(row_index, 1).font = Font(color=WHITE, bold=True)
    _table(profile_sheet, f"A4:L{len(analysis['clusters']) + 4}", "ClusterProfiles")
    _widths(profile_sheet, {"A": 10, "B": 30, "C": 13, "D": 16, "E": 17, "F": 18, "G": 17, "H": 18, "I": 16, "J": 12, "K": 14, "L": 14})
    chart = BarChart()
    chart.type = "col"
    chart.title = "Kümelere göre EV ve port yoğunluğu"
    chart.y_axis.title = "1.000 konut başına"
    chart.add_data(Reference(profile_sheet, min_col=4, max_col=5, min_row=4, max_row=len(analysis["clusters"]) + 4), titles_from_data=True)
    chart.set_categories(Reference(profile_sheet, min_col=2, min_row=5, max_row=len(analysis["clusters"]) + 4))
    chart.height = 9
    chart.width = 17
    profile_sheet.add_chart(chart, "A10")

    selection = workbook.create_sheet("K Seçimi")
    _title(selection, "Küme Sayısı Seçimi", "K=2–6 için inertia ve silhouette karşılaştırması", 7)
    _headers(selection, 4, ["K", "Inertia", "Silhouette", "Seçildi mi?"])
    for row_index, row in enumerate(analysis["kEvaluation"], 5):
        values = [row["k"], row["inertia"], row["silhouette"], "Evet" if row["k"] == analysis["selectedK"] else "Hayır"]
        for column, value in enumerate(values, 1):
            selection.cell(row_index, column, value)
    selection_end = len(analysis["kEvaluation"]) + 4
    _table(selection, f"A4:D{selection_end}", "ClusterKEvaluation")
    _widths(selection, {"A": 10, "B": 16, "C": 16, "D": 16})
    line = LineChart()
    line.title = "Silhouette score"
    line.y_axis.title = "Score"
    line.x_axis.title = "K"
    line.add_data(Reference(selection, min_col=3, min_row=4, max_row=selection_end), titles_from_data=True)
    line.set_categories(Reference(selection, min_col=1, min_row=5, max_row=selection_end))
    line.height = 9
    line.width = 16
    selection.add_chart(line, "F4")

    assignments = workbook.create_sheet("ZIP Atamaları")
    assignment_columns = [
        ("zipCode", "ZIP"), ("city", "Şehir"), ("county", "County"),
        ("clusterId", "Küme"), ("clusterLabel", "Küme adı"),
        ("vehicles", "EV"), ("publicPorts", "Kamu portu"),
        ("evPer1kHousing", "EV/1K konut"), ("portsPer1kHousing", "Port/1K konut"),
        ("medianIncome", "Medyan gelir"), ("multifamilyShare", "Çok birimli %"),
        ("workFromHomeShare", "Evden çalışma %"), ("avgCommuteMinutes", "İşe gidiş dk"),
        ("bevShare", "BEV %"),
    ]
    _headers(assignments, 1, [label for _, label in assignment_columns])
    color_by_cluster = {
        row["clusterId"]: row["color"].lstrip("#") for row in analysis["clusters"]
    }
    for row_index, row in enumerate(analysis["assignments"], 2):
        for column, (key, _) in enumerate(assignment_columns, 1):
            assignments.cell(row_index, column, row[key])
        assignments.cell(row_index, 4).fill = PatternFill("solid", fgColor=color_by_cluster[row["clusterId"]])
        assignments.cell(row_index, 4).font = Font(color=WHITE, bold=True)
    assignment_end = len(analysis["assignments"]) + 1
    _table(assignments, f"A1:N{assignment_end}", "ClusterAssignments")
    assignments.freeze_panes = "F2"
    assignments.auto_filter.ref = f"A1:N{assignment_end}"
    _widths(assignments, {"A": 11, "B": 20, "C": 18, "D": 10, "E": 28, "F": 12, "G": 14, "H": 16, "I": 17, "J": 18, "K": 17, "L": 18, "M": 15, "N": 11})

    _method_sheet(
        workbook,
        "K-Means Yöntemi",
        [
            ("Araştırma sorusu", "Washington ZIP bölgeleri EV yoğunluğu, kamu şarj portu yoğunluğu ve gelir bakımından hangi benzer profillere ayrılıyor?"),
            ("Girdiler", "EV/1.000 konut, kamu portu/1.000 konut ve medyan hane geliri."),
            ("Dönüşüm", "Yoğunluklara log(1+x) uygulandı; ardından bütün girdiler z = (x−ortalama)/standart sapma ile ölçeklendi."),
            ("Amaç fonksiyonu", "min Σ(k=1..K) Σ(xᵢ∈Cₖ) ||xᵢ−μₖ||². Her ZIP ile ait olduğu küme merkezi arasındaki kareli uzaklık küçültülür."),
            ("Silhouette", "s(i) = [b(i)−a(i)] / max[a(i),b(i)]. 1'e yaklaştıkça kümeler daha net ayrılır; 0 çevresi örtüşmeyi gösterir."),
            ("K seçimi", "K=2–6 denendi ve en yüksek silhouette değerine sahip seçenek kullanıldı."),
            ("Profil değişkenleri", "Konut tipi, evden çalışma, işe gidiş ve BEV payı modeli kurmadı; yalnız bulunan kümeleri açıklamak için raporlandı."),
            ("Sınır", "Kümeler kesin yatırım kararı değildir. Trafik, şebeke kapasitesi, maliyet ve saha uygunluğu ayrıca incelenmelidir."),
        ],
    )
    _center_workbook_cells(workbook)
    workbook.save(destination)


def export_model_workbooks(
    analysis: dict, tables: ModelTables, destination_dir: Path
) -> list[Path]:
    destination_dir.mkdir(parents=True, exist_ok=True)
    regression_path = destination_dir / "regression_analysis.xlsx"
    clustering_path = destination_dir / "clustering_analysis.xlsx"
    _regression_workbook(regression_path, analysis["regression"], tables)
    _clustering_workbook(clustering_path, analysis["clustering"], tables)
    return [regression_path, clustering_path]
